import datetime
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import db

HEADER_FILL = PatternFill('solid', fgColor='19E3B5')
HEADER_FONT = Font(bold=True, color='04130F')

# \u041f\u043e\u043d\u044f\u0442\u043d\u044b\u0435 \u043d\u0430\u0437\u0432\u0430\u043d\u0438\u044f \u043b\u0438\u0441\u0442\u043e\u0432
SHEET_TITLES = {
    'users': '\u0418\u0433\u0440\u043e\u043a\u0438',
    'roles': '\u0420\u043e\u043b\u0438',
    'tickets': '\u0411\u0438\u043b\u0435\u0442\u044b',
    'questions': '\u0412\u043e\u043f\u0440\u043e\u0441\u044b',
    'exam_attempts': '\u041f\u043e\u043f\u044b\u0442\u043a\u0438',
    'exam_answers': '\u041e\u0442\u0432\u0435\u0442\u044b',
    'anticheat_events': '\u0410\u043d\u0442\u0438\u0447\u0438\u0442',
    'violations': '\u041d\u0430\u0440\u0443\u0448\u0435\u043d\u0438\u044f',
    'servers': '\u0421\u0435\u0440\u0432\u0435\u0440\u044b',
    'news': '\u041d\u043e\u0432\u043e\u0441\u0442\u0438',
    'materials': '\u041c\u0430\u0442\u0435\u0440\u0438\u0430\u043b\u044b',
    'audit_log': '\u041b\u043e\u0433\u0438',
    'telegram_links': 'TG \u043f\u0440\u0438\u0432\u044f\u0437\u043a\u0438',
    'settings': '\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438',
}


def _style_sheet(ws, headers):
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _autosize(ws):
    for col in ws.columns:
        length = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is not None:
                length = max(length, len(str(v)))
        ws.column_dimensions[letter].width = min(48, max(10, length + 2))


def _safe_title(name, used):
    title = SHEET_TITLES.get(name, name)[:31]
    base, i = title, 2
    while title in used:
        title = (base[:28] + str(i))
        i += 1
    used.add(title)
    return title


def _access_sheet(wb):
    """\u041e\u0442\u0434\u0435\u043b\u044c\u043d\u044b\u0439 \u043b\u0438\u0441\u0442 '\u041f\u0440\u0430\u0432\u0430 \u0434\u043e\u0441\u0442\u0443\u043f\u0430' \u0438\u0437 \u0442\u0430\u0431\u043b\u0438\u0446\u044b roles."""
    try:
        roles = db.table_rows('roles')
    except Exception:
        return
    if not roles:
        return
    ws = wb.create_sheet('\u041f\u0440\u0430\u0432\u0430 \u0434\u043e\u0441\u0442\u0443\u043f\u0430')
    perm_cols = [k for k in roles[0].keys() if k.startswith('can_') or k == 'is_admin']
    headers = ['\u0420\u043e\u043b\u044c', '\u0420\u0430\u043d\u0433'] + perm_cols
    _style_sheet(ws, headers)
    for r in roles:
        row = [r.get('name'), r.get('rank')]
        for p in perm_cols:
            row.append('\u2713' if r.get(p) else '\u2014')
        ws.append(row)
    _autosize(ws)


def build_workbook(path=None):
    if path is None:
        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        # \u041a\u0440\u043e\u0441\u0441\u043f\u043b\u0430\u0442\u0444\u043e\u0440\u043c\u0435\u043d\u043d\u044b\u0439 \u043f\u0443\u0442\u044c: tempfile.gettempdir() \u0440\u0430\u0431\u043e\u0442\u0430\u0435\u0442 \u0438 \u043d\u0430 Windows, \u0438 \u043d\u0430 Linux.
        path = os.path.join(tempfile.gettempdir(), 'metrostroi_export_' + stamp + '.xlsx')
    # \u041d\u0430 \u0432\u0441\u044f\u043a\u0438\u0439 \u0441\u043b\u0443\u0447\u0430\u0439 \u0443\u0431\u0435\u0436\u0434\u0430\u0435\u043c\u0441\u044f, \u0447\u0442\u043e \u043f\u0430\u043f\u043a\u0430 \u0441\u0443\u0449\u0435\u0441\u0442\u0432\u0443\u0435\u0442.
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    # \u0421\u043d\u0430\u0447\u0430\u043b\u0430 \u0441\u0432\u043e\u0434\u043d\u044b\u0439 \u043b\u0438\u0441\u0442 \u043f\u0440\u0430\u0432
    _access_sheet(wb)
    used.add('\u041f\u0440\u0430\u0432\u0430 \u0434\u043e\u0441\u0442\u0443\u043f\u0430')
    for table in db.list_tables():
        try:
            rows = db.table_rows(table)
        except Exception:
            continue
        ws = wb.create_sheet(_safe_title(table, used))
        if rows:
            headers = list(rows[0].keys())
            _style_sheet(ws, headers)
            for r in rows:
                ws.append([_cell(r.get(h)) for h in headers])
        else:
            ws.cell(row=1, column=1, value='(\u043f\u0443\u0441\u0442\u043e)')
        _autosize(ws)
    wb.save(path)
    return path


def _cell(v):
    if isinstance(v, (dict, list)):
        import json
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, (bytes, bytearray)):
        try:
            return v.decode('utf-8', 'ignore')
        except Exception:
            return str(v)
    return v
